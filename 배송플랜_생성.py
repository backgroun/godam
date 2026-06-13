"""
배송플랜 자동화 스크립트
사용법: python 배송플랜_생성.py [발주파일.xlsx] [거래처마스터.xlsx]
기본값: 스크립트와 같은 폴더에 있는 파일 자동 탐색
"""

import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import sys
import os
import glob

# ============================================================
# ★ 설정 영역 - 여기만 수정하면 됩니다
# ============================================================

# 기사님 이름 목록 (순서대로 구역 배정)
DRIVERS = [
    "기사1",
    "기사2",
    "기사3",
    "기사4",
]

# 구역 정의: {구역명: [주소에 포함될 키워드 목록]}
# 주소의 시/구 단위와 매칭됩니다
ZONES = {
    "강남권":    ["강남구", "서초구", "송파구", "강동구"],
    "강북권":    ["노원구", "은평구", "동대문구", "성북구", "중계"],
    "서울중부권": ["마포구", "영등포구", "금천구", "구로구", "동작구", "중구", "용산구"],
    "서울서부권": ["강서구", "양천구"],
    "경기동북권": ["양주시", "의정부시", "동두천시", "연천군", "포천시"],
    "경기동부권": ["구리시", "남양주시", "가평군"],
    "경기남부권": ["성남시", "분당구", "용인시", "화성시", "안성시", "평택시", "수원시", "오산시"],
    "경기서부권": ["김포시", "파주시", "시흥시", "부천시", "광명시"],
    "인천권":    ["인천"],
    "원거리":   ["구미시", "부산", "울산", "전주", "정읍", "광주", "대구", "대전"],
}

# 기사별 담당 구역 배정
# {기사이름: [담당 구역명 목록]}
DRIVER_ZONES = {
    "기사1": ["강남권", "서울중부권"],
    "기사2": ["강북권", "서울서부권", "경기동북권"],
    "기사3": ["경기동부권", "경기남부권"],
    "기사4": ["경기서부권", "인천권", "원거리"],
}

# ============================================================
# 주소 정규화 함수
# ============================================================

def normalize_address(addr):
    """주소에서 시/구 단위를 표준화해서 추출"""
    if pd.isna(addr):
        return None, None
    addr = str(addr).strip()
    parts = addr.split()
    if len(parts) < 2:
        return parts[0] if parts else None, None

    # 시도 정규화 (서울시/서울특별시/서울 → 서울)
    sido_map = {
        "서울시": "서울", "서울특별시": "서울", "서울을시": "서울",
        "경기도": "경기", "경기": "경기",
        "인천시": "인천", "인천광역시": "인천",
        "부산시": "부산", "부산광역시": "부산",
        "대구시": "대구", "대구광역시": "대구",
        "광주시": "광주", "광주광역시": "광주",
        "대전시": "대전", "대전광역시": "대전",
        "울산시": "울산", "울산광역시": "울산",
        "경북": "경북", "경상북도": "경북",
        "경남": "경남", "경상남도": "경남",
        "전북": "전북", "전라북도": "전북",
        "전남": "전남", "전라남도": "전남",
        "충북": "충북", "충청북도": "충북",
        "충남": "충남", "충청남도": "충남",
        "강원": "강원", "강원도": "강원", "강원특별자치도": "강원",
        "세종특별자치시": "세종",
        "제주특별자치도": "제주", "제주도": "제주",
    }

    sido = sido_map.get(parts[0], parts[0])
    sigungu = parts[1] if len(parts) > 1 else ""
    return sido, sigungu


def get_zone(addr):
    """주소로부터 구역명 반환"""
    if pd.isna(addr):
        return "주소없음"
    addr_str = str(addr)
    for zone_name, keywords in ZONES.items():
        for kw in keywords:
            if kw in addr_str:
                return zone_name
    return "기타"


def get_driver(zone):
    """구역으로부터 담당 기사 반환"""
    for driver, zones in DRIVER_ZONES.items():
        if zone in zones:
            return driver
    return "미배정"


# ============================================================
# 메인 처리
# ============================================================

def find_file(pattern_list, directory="."):
    """파일 자동 탐색"""
    for pattern in pattern_list:
        files = glob.glob(os.path.join(directory, pattern))
        if files:
            return sorted(files)[-1]  # 가장 최근 파일
    return None


def load_master(master_path):
    df = pd.read_excel(master_path)
    df.columns = df.columns.str.strip()
    # 컬럼명 유연하게 처리
    col_map = {}
    for c in df.columns:
        if '거래처' in c and '명' in c or '거래처명' in c:
            col_map[c] = '거래처명'
        elif '정제주소' in c or ('구글' in c):
            col_map[c] = '주소'
        elif '주소' in c and '정제' not in c:
            col_map[c] = '원주소'
        elif '코' in c and '드' in c or c.strip() == '코드':
            col_map[c] = '코드'
    df = df.rename(columns=col_map)
    return df


def load_orders(order_path):
    df = pd.read_excel(order_path)
    df = df.dropna(subset=['거래처명'])
    # 빈 행 제거 (거래처명이 NaN인 행)
    df = df[df['거래처명'].astype(str).str.strip() != 'nan']
    return df


def build_plan(orders_df, master_df):
    """발주 + 마스터 조인 후 구역/기사 배정"""
    # 주소 컬럼 선택 (정제주소 우선, 없으면 원주소)
    addr_col = '주소' if '주소' in master_df.columns else '원주소'

    merged = orders_df.merge(
        master_df[['거래처명', addr_col]].drop_duplicates('거래처명'),
        on='거래처명',
        how='left'
    )
    merged = merged.rename(columns={addr_col: '배송주소'})

    # 구역 및 기사 배정
    merged['구역'] = merged['배송주소'].apply(get_zone)
    merged['담당기사'] = merged['구역'].apply(get_driver)

    # 시도/시군구 분리
    parsed = merged['배송주소'].apply(lambda x: normalize_address(x))
    merged['시도'] = [p[0] for p in parsed]
    merged['시군구'] = [p[1] for p in parsed]

    return merged


# ============================================================
# Excel 출력
# ============================================================

ZONE_COLORS = {
    "강남권":    "FFD700",
    "강북권":    "98FB98",
    "서울중부권": "87CEEB",
    "서울서부권": "DDA0DD",
    "경기동북권": "FFA07A",
    "경기동부권": "F0E68C",
    "경기남부권": "AFEEEE",
    "경기서부권": "FFB6C1",
    "인천권":   "E0E0E0",
    "원거리":   "C0C0C0",
    "기타":     "FFFFFF",
    "주소없음":  "FFCCCC",
}

DRIVER_COLORS = {
    "기사1": "FFF2CC",
    "기사2": "E2EFDA",
    "기사3": "DAEEF3",
    "기사4": "FCE4D6",
}

def thin_border():
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

def write_excel(plan_df, output_path, order_date):
    wb = openpyxl.Workbook()

    # ── 시트1: 기사별 배송플랜 ──
    ws1 = wb.active
    ws1.title = "배송플랜"

    title_font = Font(name="맑은 고딕", bold=True, size=14)
    header_font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    body_font = Font(name="맑은 고딕", size=10)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 제목
    ws1.merge_cells("A1:J1")
    ws1["A1"] = f"배송플랜 — {order_date}"
    ws1["A1"].font = Font(name="맑은 고딕", bold=True, size=14)
    ws1["A1"].alignment = center
    ws1.row_dimensions[1].height = 30

    # 기사별로 섹션 분리 출력
    row = 3
    driver_order = list(DRIVER_ZONES.keys()) + ["미배정"]

    for driver in driver_order:
        subset = plan_df[plan_df['담당기사'] == driver]
        if subset.empty:
            continue

        drv_color = DRIVER_COLORS.get(driver, "EEEEEE")

        # 기사 헤더
        ws1.merge_cells(f"A{row}:J{row}")
        ws1[f"A{row}"] = f"▶ {driver}  ({len(subset['거래처명'].unique())}개 거래처 / {subset['수량'].sum():.0f}건)"
        ws1[f"A{row}"].font = Font(name="맑은 고딕", bold=True, size=11)
        ws1[f"A{row}"].fill = PatternFill("solid", fgColor="4472C4")
        ws1[f"A{row}"].font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
        ws1[f"A{row}"].alignment = left
        ws1.row_dimensions[row].height = 22
        row += 1

        # 컬럼 헤더
        headers = ["순번", "구역", "거래처명", "배송주소", "품목명", "수량", "구분", "비고1", "비고2", "납기일"]
        for ci, h in enumerate(headers, 1):
            cell = ws1.cell(row=row, column=ci, value=h)
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="595959")
            cell.alignment = center
            cell.border = thin_border()
        ws1.row_dimensions[row].height = 18
        row += 1

        # 구역별 정렬
        subset_sorted = subset.sort_values(['구역', '거래처명'])
        seq = 1
        prev_zone = None

        for _, r in subset_sorted.iterrows():
            zone = r.get('구역', '기타')
            if zone != prev_zone:
                prev_zone = zone
                seq_in_zone = 1

            zone_color = ZONE_COLORS.get(zone, "FFFFFF")
            values = [
                seq,
                zone,
                r.get('거래처명', ''),
                r.get('배송주소', ''),
                r.get('품목명', ''),
                r.get('수량', ''),
                r.get('구분', ''),
                r.get('비고1', ''),
                r.get('비고2', ''),
                str(r.get('납기일', ''))[:10] if pd.notna(r.get('납기일')) else '',
            ]
            for ci, v in enumerate(values, 1):
                cell = ws1.cell(row=row, column=ci, value=v)
                cell.font = body_font
                cell.alignment = center if ci in [1, 2, 6, 7, 10] else left
                cell.border = thin_border()
                if ci == 2:
                    cell.fill = PatternFill("solid", fgColor=zone_color)
                elif ci in [3, 4]:
                    cell.fill = PatternFill("solid", fgColor=drv_color)
            ws1.row_dimensions[row].height = 18
            row += 1
            seq += 1

        row += 1  # 기사 섹션 간격

    # 열 너비 조정
    col_widths = [6, 12, 22, 45, 30, 7, 8, 25, 25, 12]
    for ci, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── 시트2: 미매칭 거래처 ──
    ws2 = wb.create_sheet("주소없음_확인필요")
    no_addr = plan_df[plan_df['배송주소'].isna() | (plan_df['구역'] == '주소없음')]

    ws2["A1"] = "주소 없음 — 확인 필요"
    ws2["A1"].font = Font(name="맑은 고딕", bold=True, size=12, color="FF0000")
    ws2.merge_cells("A1:E1")

    headers2 = ["거래처명", "품목명", "수량", "구분", "비고1"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=2, column=ci, value=h)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="C00000")
        cell.alignment = center
        cell.border = thin_border()

    for ri, (_, r) in enumerate(no_addr.iterrows(), 3):
        for ci, col in enumerate(headers2, 1):
            cell = ws2.cell(row=ri, column=ci, value=r.get(col, ''))
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = thin_border()

    # ── 시트3: 구역별 요약 ──
    ws3 = wb.create_sheet("구역별요약")
    ws3["A1"] = "구역별 요약"
    ws3["A1"].font = Font(name="맑은 고딕", bold=True, size=12)
    ws3.merge_cells("A1:D1")

    headers3 = ["구역", "담당기사", "거래처 수", "총 수량"]
    for ci, h in enumerate(headers3, 1):
        cell = ws3.cell(row=2, column=ci, value=h)
        cell.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="404040")
        cell.alignment = center
        cell.border = thin_border()

    summary = (plan_df
               .groupby(['구역', '담당기사'])
               .agg(거래처수=('거래처명', 'nunique'), 총수량=('수량', 'sum'))
               .reset_index()
               .sort_values('구역'))

    for ri, (_, r) in enumerate(summary.iterrows(), 3):
        zone = r['구역']
        zcolor = ZONE_COLORS.get(zone, "FFFFFF")
        vals = [zone, r['담당기사'], r['거래처수'], r['총수량']]
        for ci, v in enumerate(vals, 1):
            cell = ws3.cell(row=ri, column=ci, value=v)
            cell.font = Font(name="맑은 고딕", size=10)
            cell.border = thin_border()
            cell.alignment = center
            if ci == 1:
                cell.fill = PatternFill("solid", fgColor=zcolor)

    for ci, w in enumerate([15, 12, 12, 12], 1):
        ws3.column_dimensions[get_column_letter(ci)].width = w

    wb.save(output_path)
    print(f"✅ 저장 완료: {output_path}")


# ============================================================
# 실행 진입점
# ============================================================

def main():
    script_dir = os.path.dirname(os.path.abspath(__file__))

    # 파일 경로 지정 (인수 없으면 자동 탐색)
    if len(sys.argv) >= 3:
        order_path = sys.argv[1]
        master_path = sys.argv[2]
    else:
        order_path = find_file(["meatpath_진행주문_*.xlsx", "발주_*.xlsx", "*진행주문*.xlsx"], script_dir)
        master_path = find_file(["upload_고담_정제.xlsx", "*거래처*.xlsx", "*고담*.xlsx"], script_dir)

        if not order_path or not master_path:
            print("❌ 파일을 찾을 수 없습니다.")
            print("   사용법: python 배송플랜_생성.py [발주파일.xlsx] [거래처마스터.xlsx]")
            sys.exit(1)

    print(f"📄 발주 파일: {order_path}")
    print(f"📄 마스터 파일: {master_path}")

    orders_df = load_orders(order_path)
    master_df = load_master(master_path)

    plan_df = build_plan(orders_df, master_df)

    # 날짜 추출
    order_date = str(orders_df['납기일'].dropna().iloc[0])[:10] if '납기일' in orders_df.columns else datetime.today().strftime('%Y-%m-%d')

    output_path = os.path.join(script_dir, f"배송플랜_{order_date.replace('-','')}.xlsx")
    write_excel(plan_df, output_path, order_date)

    # 콘솔 요약
    print()
    print("=== 배정 요약 ===")
    summary = plan_df.groupby('담당기사').agg(
        거래처수=('거래처명', 'nunique'),
        총수량=('수량', 'sum')
    )
    print(summary.to_string())

    no_addr_cnt = plan_df['배송주소'].isna().sum()
    if no_addr_cnt > 0:
        print(f"\n⚠️  주소 없는 행 {no_addr_cnt}건 → '주소없음_확인필요' 시트 확인")


if __name__ == "__main__":
    main()
